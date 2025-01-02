#!/usr/bin/env python3
"""
Merges the output of sabayomikun and Katikati Counter.
"""
import sys
import os
from os.path import expanduser
import datetime
import csv
import linecache
import shutil
import re
try:
    import openpyxl
except ModuleNotFoundError:
    print('openpyxl is not installed.')
    print('pip install openpyxl')
    sys.exit()
if sys.version_info.major == 2:
    import Tkinter
    import tkMessageBox
    import tkFileDialog
    TKROOT = Tkinter.Tk()
elif sys.version_info.major == 3:
    import tkinter
    from tkinter import messagebox as tkMessageBox
    from tkinter import filedialog as tkFileDialog
    TKROOT = tkinter.Tk()
TKROOT.withdraw()
if os.name == 'nt':
    IDIR = expanduser("~") + '/Desktop'
else:
    IDIR = os.environ['HOME'] + '/Desktop'
FTYP1 = [('From sabayomikun', '*.csv')]
COMBINEMODE = True
INPUTDIRFLAG0 = tkMessageBox.askquestion('katisaba', 'Do you combine auto-counted foci\nand manually-counted numbers?')
if INPUTDIRFLAG0 == str('no'):
    COMBINEMODE = False
else:
    if not tkMessageBox.askokcancel('katisaba', 'Select automatic count csv file.\n(The output file of sabayomikun)'):
        sys.exit()
    CSVPATH1 = tkFileDialog.askopenfilename(filetypes=FTYP1, initialdir=IDIR)
    if CSVPATH1 == '':
        sys.exit()
    if not os.path.exists(CSVPATH1):
        tkMessageBox.showinfo('katisaba', 'Missing input file.')
        sys.exit()
    if not os.path.isfile(CSVPATH1):
        tkMessageBox.showinfo('katisaba', 'Missing input file.')
        sys.exit()
    CSVFILE1 = open(CSVPATH1, mode='r')
    CSVTEXT1 = CSVFILE1.readlines()
    CSVFILE1.close()
    BASENAME1 = os.path.basename(CSVPATH1)
    DIRNAME1 = os.path.dirname(CSVPATH1)
    ROOTNAME1, EXT1 = os.path.splitext(BASENAME1)
    LINESVERT = 1
    while LINESVERT < 9:
        INDEXLINE1 = linecache.getline(CSVPATH1, LINESVERT).replace('\n', '').split(',')
        for uf in INDEXLINE1:
            if not uf.isdigit():
                if not uf == 'ND':
                    tkMessageBox.showinfo('katisaba', 'Inpt file format does not match.\nIs this the output file of sabayomikun?')
                    sys.exit()
        if len(INDEXLINE1) != 12:
            tkMessageBox.showinfo('katisaba', 'Inpt file format does not match.\nIs this the output file of sabayomikun?')
            sys.exit()
        LINESVERT += 1
FTYP2 = [('From Katikati Counter', '*.csv')]
if not tkMessageBox.askokcancel('katisaba', 'Select manual count csv file.\n(The output file of Katikati Counter)'):
    sys.exit()
CSVPATH2 = tkFileDialog.askopenfilename(filetypes=FTYP2, initialdir=IDIR)
if CSVPATH2 == '':
    sys.exit()
if not os.path.exists(CSVPATH2):
    tkMessageBox.showinfo('katisaba', 'Missing input file.')
    sys.exit()
if not os.path.isfile(CSVPATH2):
    tkMessageBox.showinfo('katisaba', 'Missing input file.')
    sys.exit()
CSVFILE2 = open(CSVPATH2, mode='r')
CSVTEXT2 = CSVFILE2.readlines()
INDEXLINE2 = linecache.getline(CSVPATH2, 1).replace('\n', '').split(',')
INDEXLINE3 = CSVTEXT2[1:]
CSVFILE2.close()
if 'Red' not in INDEXLINE2 or 'Green' not in INDEXLINE2 or 'Blue' not in INDEXLINE2:
    tkMessageBox.showinfo('katisaba', 'Index line is modified, or this is not\nthe output file of Katikati Counter.')
    sys.exit()
REDINDEX = INDEXLINE2.index('Red')
GREENINDEX = INDEXLINE2.index('Green')
BLUEINDEX = INDEXLINE2.index('Blue')
if not 0 < REDINDEX < 4 or not 0 < GREENINDEX < 4 or not 0 < BLUEINDEX < 4:
    tkMessageBox.showinfo('katisaba', 'Index line is modified, or this is not\nthe output file of Katikati Counter.')
    sys.exit()
for lines in INDEXLINE3:
    line = lines.replace('\n', '').split(',')
    if not re.search("-detected[1-2]$", line[0]):
        if not re.match("^[A-H][1-9]$", line[0]):
            if not re.match("^[A-H]1[0-2]$", line[0]):
                tkMessageBox.showinfo('katisaba', 'This is not the correct output file of\nKatikati Counter, or you included\nnon-graphic files.')
                sys.exit()
    if not line[1].isdigit() or not line[2].isdigit() or not line[3].isdigit():
        tkMessageBox.showinfo('katisaba', 'This is not the output file of\nKatikati Counter.')
        sys.exit()
TKROOT.update()
BASENAME2 = os.path.basename(CSVPATH2)
DIRNAME2 = os.path.dirname(CSVPATH2)
ROOTNAME2, EXT2 = os.path.splitext(BASENAME2)
TKROOT.update()
VWELL = [chr(i) for i in range(65, 65 + 8)]
VWELLE = [chr(i) for i in range(65, 65 + 12)]
HWELL = (range(1, 13))
DATETIME = datetime.datetime.today().strftime("%Y_%m_%d_%H_%M_%S")
AUTOCOUNT_EXCL = {}
for vwellnm in VWELL:
    vwellloc = int(VWELL.index(vwellnm)) + 1
    for hwellnm in HWELL:
        hwellloc = VWELLE[int(hwellnm) - 1]
        AUTOCOUNT_EXCL[(vwellnm + str(hwellnm))] = hwellloc + str(vwellloc)
if COMBINEMODE:
    AUTOCOUNT_AUTO_ARR = []
    CSVDATA1 = csv.reader(CSVTEXT1)
    for row in CSVDATA1:
        AUTOCOUNT_AUTO_ARR.append(row)
    AUTOCOUNT_AUTO = {}
    for vwellnm in VWELL:
        for hwellnm in HWELL:
            AUTOCOUNT_AUTO[(vwellnm + str(hwellnm))] = AUTOCOUNT_AUTO_ARR[VWELL.index(vwellnm)][HWELL.index(hwellnm)]
HANDCOUNT = []
for dm in CSVTEXT2:
    if 'Red' not in dm:
        HANDCOUNT.append(dm.replace('\n', ''))
HANDCOUNT_R = {}
HANDCOUNT_B = {}
HANDCOUNT_G = {}
for lm in HANDCOUNT:
    lm = lm.split(',')
    HANDCOUNT_R[lm[0]] = lm[REDINDEX]
    HANDCOUNT_B[lm[0]] = lm[BLUEINDEX]
    HANDCOUNT_G[lm[0]] = lm[GREENINDEX]
HANDCOUNT_FIN = []
HANDCOUNT_RED = []
HANDCOUNT_BLUE = []
HANDCOUNT_GREEN = []
CAUTION = []
if COMBINEMODE:
    HANDCOUNT_ORIG = []
for vwellnm in VWELL:
    for hwellnm in HWELL:
        callindex = vwellnm + str(hwellnm)
        temparr = []
        try:
            temparr.append(HANDCOUNT_R[callindex])
        except KeyError:
            tkMessageBox.showinfo('katisaba', 'Some of count records are missing.')
            sys.exit()
        temparr.append(HANDCOUNT_B[callindex])
        temparr.append(HANDCOUNT_G[callindex])
        if COMBINEMODE:
            temparr.append(AUTOCOUNT_AUTO[callindex])
            if temparr[0] == '' or temparr[1] == '' or temparr[2] == '':
                newautodata = ''
            elif temparr[3] == 'ND':
                newautodata = 'ND'
            else:
                newautodata = int(temparr[3]) + int(temparr[0]) - int(temparr[1])
            if newautodata == 'ND':
                HANDCOUNT_ORIG.append('ND')
            else:
                HANDCOUNT_ORIG.append(int(temparr[3]))
        else:
            if temparr[0] == '' or temparr[1] == '':
                newautodata = 'ND'
            else:
                newautodata = int(temparr[0]) - int(temparr[1])
        HANDCOUNT_FIN.append(newautodata)
        if temparr[0] == '':
            HANDCOUNT_RED.append('ND')
        else:
            HANDCOUNT_RED.append(temparr[0])
        if temparr[1] == '':
            HANDCOUNT_BLUE.append('ND')
        else:
            HANDCOUNT_BLUE.append(temparr[1])
        if temparr[2] == '':
            HANDCOUNT_GREEN.append('ND')
        else:
            HANDCOUNT_GREEN.append(temparr[2])
        if int(temparr[2]) > 0:
            CAUTION.append(AUTOCOUNT_EXCL[callindex])
HANDCOUNT_FIN = [HANDCOUNT_FIN[x:x + 12] for x in range(0, len(HANDCOUNT_FIN), 12)]
HANDCOUNT_RED = [HANDCOUNT_RED[x:x + 12] for x in range(0, len(HANDCOUNT_RED), 12)]
HANDCOUNT_BLUE = [HANDCOUNT_BLUE[x:x + 12] for x in range(0, len(HANDCOUNT_BLUE), 12)]
HANDCOUNT_GREEN = [HANDCOUNT_GREEN[x:x + 12] for x in range(0, len(HANDCOUNT_GREEN), 12)]
if COMBINEMODE:
    HANDCOUNT_ORIG = [HANDCOUNT_ORIG[x:x + 12] for x in range(0, len(HANDCOUNT_ORIG), 12)]
WB = openpyxl.Workbook()
WS = WB.active
WS.title = 'Summary'
for row in HANDCOUNT_FIN:
    WS.append(row)
for cautionwell in CAUTION:
    WS[cautionwell].font = openpyxl.styles.fonts.Font(color='00FF00')
if COMBINEMODE:
    WSORIG = WB.create_sheet('Original')
    for row in HANDCOUNT_ORIG:
        WSORIG.append(row)
WSRED = WB.create_sheet('Red(appended)')
for row in HANDCOUNT_RED:
    WSRED.append(row)
WSBLUE = WB.create_sheet('Blue(removed)')
for row in HANDCOUNT_BLUE:
    WSBLUE.append(row)
WSGREEN = WB.create_sheet('Green(caution)')
for row in HANDCOUNT_GREEN:
    WSGREEN.append(row)
if os.path.exists(DIRNAME2 + '/' + ROOTNAME2 + '.xlsx'):
    shutil.move(DIRNAME2 + '/' + ROOTNAME2 + '.xlsx', DIRNAME2 + '/' + ROOTNAME2 + '_old_' + DATETIME + '.xlsx')
WB.save(DIRNAME2 + '/' + ROOTNAME2 + '.xlsx')
tkMessageBox.showinfo('katisaba', 'Task finished.')
